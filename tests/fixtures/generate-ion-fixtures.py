"""Regenerate deterministic synthetic workbook and browser E2E fixtures."""

from datetime import datetime
from io import BytesIO
from pathlib import Path
import re
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

from openpyxl import Workbook


OUTPUT = Path(__file__).parent
FIXED_DOCUMENT_TIME = datetime(2000, 1, 1, 0, 0, 0)
FIXED_ZIP_TIME = (1980, 1, 1, 0, 0, 0)


def save_deterministic(workbook: Workbook, destination: Path) -> None:
    """Write stable XLSX bytes by fixing document and ZIP member timestamps."""
    workbook.properties.created = FIXED_DOCUMENT_TIME
    workbook.properties.modified = FIXED_DOCUMENT_TIME
    source_buffer = BytesIO()
    workbook.save(source_buffer)

    output_buffer = BytesIO()
    with ZipFile(BytesIO(source_buffer.getvalue()), "r") as source:
        with ZipFile(output_buffer, "w", compression=ZIP_DEFLATED, compresslevel=9) as target:
            for name in sorted(source.namelist()):
                original = source.getinfo(name)
                payload = source.read(name)
                if name == "docProps/core.xml":
                    payload = re.sub(
                        rb"<dcterms:modified[^>]*>.*?</dcterms:modified>",
                        b'<dcterms:modified xmlns:dcterms="http://purl.org/dc/terms/" '
                        b'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" '
                        b'xsi:type="dcterms:W3CDTF">2000-01-01T00:00:00Z</dcterms:modified>',
                        payload,
                    )
                stable = ZipInfo(name, FIXED_ZIP_TIME)
                stable.compress_type = ZIP_DEFLATED
                stable.external_attr = original.external_attr
                stable.create_system = 0
                target.writestr(stable, payload)

    destination.write_bytes(output_buffer.getvalue())


def preferred_sheet_fixture() -> None:
    workbook = Workbook()
    workbook.active.title = "说明"
    workbook.active.append(["not data"])
    sheet = workbook.create_sheet("站点数据")
    sheet.append(["2024 年水溶性离子"])
    sheet.append(["来源", "test fixture"])
    sheet.append(["日期时间", "NO₃⁻", "SO₄²⁻", "NH₄⁺"])
    sheet.append(["单位", "μg/m³", "ug/m3", "ug·m-3"])
    sheet.append([datetime(2024, 1, 2, 3), 0, 2.5, 1])
    save_deterministic(workbook, OUTPUT / "ions-small.xlsx")


def fallback_sheet_fixture() -> None:
    workbook = Workbook()
    workbook.active.title = "说明"
    workbook.active.append(["not data"])
    sheet = workbook.create_sheet("有效数据")
    sheet.append(["time", "NO3-", "SO4", "NH4+"])
    sheet.append(["2024-01-01 00:00", 1, 2, 3])
    save_deterministic(workbook, OUTPUT / "ions-fallback.xlsx")


def e2e_workflow_fixture() -> None:
    workbook = Workbook()
    workbook.active.title = "站点数据"
    workbook.active.append(["日期时间", "NO3-", "SO4", "NH4+"])
    workbook.active.append([datetime(2024, 11, 1, 0), 8.5, 3.2, 5.1])
    save_deterministic(workbook, OUTPUT / "ions-e2e.xlsx")


def user_data_fixtures() -> None:
    csv_text = (
        "\ufefftimestamp,Temperature,Tracer\r\n"
        "2024-11-01 00:00:00,-5,1\r\n"
        "2024-11-01 01:00:00,-4,-2\r\n"
        "2024-11-01 02:00:00,-3,3\r\n"
    )
    (OUTPUT / "user-data-small.csv").write_bytes(csv_text.encode("utf-8"))

    workbook = Workbook()
    workbook.active.title = "Measurements"
    workbook.active.append(["time", "timestamp", "Temperature", "Tracer"])
    workbook.active.append([datetime(2024, 11, 1, 0), datetime(2024, 11, 1, 0), -5, 1])
    workbook.active.append([datetime(2024, 11, 1, 1), datetime(2024, 11, 1, 1), -4, -2])
    workbook.active.append([datetime(2024, 11, 1, 2), datetime(2024, 11, 1, 2), -3, 3])
    save_deterministic(workbook, OUTPUT / "user-data-small.xlsx")


if __name__ == "__main__":
    preferred_sheet_fixture()
    fallback_sheet_fixture()
    e2e_workflow_fixture()
    user_data_fixtures()
